import concurrent.futures
import os
from contextlib import contextmanager
from typing import Any, Callable, Dict, Iterable, Iterator, List, Optional, Sequence, Tuple

import openpyxl


DEFAULT_EXCEL_EXTENSIONS = (".xlsx", ".xls")


def safe_to_str(value: Any, strip: bool = True) -> str:
    if value is None:
        return ""
    text = str(value)
    return text.strip() if strip else text


def is_blank_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def build_combined_key(
    key_value: Any,
    match_value: Any,
    separator: str = "|",
) -> Optional[str]:
    key_text = safe_to_str(key_value, strip=True)
    match_text = safe_to_str(match_value, strip=True)
    if not key_text or not match_text:
        return None
    return f"{key_text}{separator}{match_text}"


def iter_excel_files(
    folder_path: str,
    extensions: Sequence[str] = DEFAULT_EXCEL_EXTENSIONS,
    include_temp_files: bool = True,
    case_sensitive: bool = False,
) -> List[str]:
    file_paths: List[str] = []
    normalized_ext = tuple(ext.lower() for ext in extensions)
    for root, _, files in os.walk(folder_path):
        for file_name in files:
            if case_sensitive:
                matches_ext = file_name.endswith(tuple(extensions))
            else:
                matches_ext = file_name.lower().endswith(normalized_ext)
            if not matches_ext:
                continue
            if not include_temp_files and file_name.startswith("~$"):
                continue
            file_paths.append(os.path.join(root, file_name))
    return file_paths


@contextmanager
def open_workbook(
    file_path: str,
    read_only: bool = False,
    data_only: bool = False,
) -> Iterator[openpyxl.Workbook]:
    workbook = None
    try:
        workbook = openpyxl.load_workbook(
            filename=file_path,
            read_only=read_only,
            data_only=data_only,
        )
        yield workbook
    finally:
        if workbook is not None:
            workbook.close()


def apply_cell_updates(file_path: str, updates: Dict[Tuple[int, int], Any]) -> bool:
    if not updates:
        return True

    workbook = None
    try:
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        for (row_idx, col_idx), value in updates.items():
            worksheet.cell(row=row_idx, column=col_idx).value = value
        workbook.save(file_path)
        return True
    except Exception:
        return False
    finally:
        if workbook is not None:
            workbook.close()


def run_parallel_map(
    items: Sequence[Any],
    worker: Callable[[Any], Any],
    max_workers_cap: int = 32,
) -> List[Any]:
    if not items:
        return []

    max_workers = min(max_workers_cap, len(items))
    max_workers = max(1, max_workers)

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_index = {
            executor.submit(worker, item): index for index, item in enumerate(items)
        }
        results: List[Any] = [None] * len(items)
        for future in concurrent.futures.as_completed(future_to_index):
            index = future_to_index[future]
            results[index] = future.result()
        return results


def run_parallel_sum(
    items: Sequence[Any],
    worker: Callable[[Any], int],
    max_workers_cap: int = 32,
) -> int:
    return sum(run_parallel_map(items, worker, max_workers_cap=max_workers_cap))
