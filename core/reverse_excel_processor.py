import pandas as pd
import os
import concurrent.futures
import openpyxl
import time

class ReverseExcelProcessor:
    def __init__(self, log_callback=None):
        self.master_file_path = ""
        self.target_folder = ""
        self.log_callback = log_callback or (lambda msg: None)
        
        # Column indices (0-based)
        self.target_key_col = 0
        self.target_match_col = 1
        self.target_content_col = 2

        self.master_key_col = 1
        self.master_match_col = 2
        self.master_update_col = 3

    def set_target_columns(self, key_col, match_col, content_col):
        self.target_key_col = key_col
        self.target_match_col = match_col
        self.target_content_col = content_col

    def set_master_columns(self, key_col, match_col, update_col):
        self.master_key_col = key_col
        self.master_match_col = match_col
        self.master_update_col = update_col

    def set_master_file(self, file_path):
        self.master_file_path = file_path

    def set_target_folder(self, folder_path):
        self.target_folder = folder_path

    def log(self, message):
        self.log_callback(message)

    def process_files(self):
        if not self.master_file_path or not self.target_folder:
            raise ValueError("请先选择 Master 文件和目标文件夹！")

        start_time = time.time()
        self.log("开始处理文件...")

        # 1. 从小文件中读取数据并构建数据字典
        self.log("正在从目标文件夹中的小文件读取数据...")
        target_data_dict = self._read_target_files()
        self.log(f"从小文件中读取了 {len(target_data_dict)} 条数据。")

        # 2. 更新Master文件
        self.log("正在更新 Master 文件...")
        updated_count = self._update_master_file(target_data_dict)
        self.log(f"Master 文件更新完成，共更新 {updated_count} 处数据。")

        total_time = time.time() - start_time
        self.log(f"总耗时: {total_time:.2f}秒")

        return updated_count

    def _read_target_files(self):
        file_paths = []
        for root, _, files in os.walk(self.target_folder):
            file_paths.extend(
                os.path.join(root, file)
                for file in files
                if file.lower().endswith(('.xlsx', '.xls'))
            )

        self.log(f"找到 {len(file_paths)} 个目标文件")
        
        data_dict = {}
        with concurrent.futures.ThreadPoolExecutor() as executor:
            futures = [executor.submit(self._read_single_target_file, fp) for fp in file_paths]
            for future in concurrent.futures.as_completed(futures):
                data_dict.update(future.result())
        
        return data_dict

    def _read_single_target_file(self, file_path):
        local_dict = {}
        try:
            wb = openpyxl.load_workbook(filename=file_path, read_only=True)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2):
                try:
                    # Ensure row has enough columns to avoid IndexError
                    if len(row) <= max(self.target_key_col, self.target_match_col, self.target_content_col):
                        continue

                    key_cell = row[self.target_key_col]
                    match_cell = row[self.target_match_col]
                    content_cell = row[self.target_content_col]

                    key = str(key_cell.value).strip() if key_cell.value is not None else ''
                    match_val = str(match_cell.value).strip() if match_cell.value is not None else ''
                    content_val = str(content_cell.value) if content_cell.value is not None else ''

                    if key and match_val:
                        combined_key = f"{key}|{match_val}"
                        local_dict[combined_key] = content_val
                except IndexError:
                    # This is a safeguard, though the length check should prevent it.
                    continue
        except Exception as e:
            self.log(f"读取文件 {os.path.basename(file_path)} 时出错: {e}")
        return local_dict

    def _update_master_file(self, target_data_dict):
        if not target_data_dict:
            self.log("没有从目标文件中读取到任何数据，无需更新。")
            return 0

        updated_count = 0
        try:
            # Load the workbook with openpyxl to preserve styles
            wb = openpyxl.load_workbook(self.master_file_path)
            ws = wb.active

            # Iterate over rows in the worksheet
            for row in ws.iter_rows():
                # Ensure row has enough columns to avoid IndexError
                if len(row) <= max(self.master_key_col, self.master_match_col, self.master_update_col):
                    continue

                master_key_cell = row[self.master_key_col]
                master_match_cell = row[self.master_match_col]

                master_key = str(master_key_cell.value).strip() if master_key_cell.value is not None else ''
                master_match_value = str(master_match_cell.value).strip() if master_match_cell.value is not None else ''

                if not master_key or not master_match_value:
                    continue

                combined_key = f"{master_key}|{master_match_value}"
                if combined_key in target_data_dict:
                    # Update the cell value in the update column
                    update_cell = row[self.master_update_col]
                    update_cell.value = target_data_dict[combined_key]
                    updated_count += 1

            if updated_count > 0:
                # Save the workbook, which preserves styles
                wb.save(self.master_file_path)
                
        except Exception as e:
            self.log(f"更新 Master 文件时出错: {e}")
            return 0
        return updated_count