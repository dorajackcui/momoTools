import argparse
import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, List, Tuple

import openpyxl


@dataclass
class DiffItem:
    row: int
    col: int
    expected: Any
    actual: Any


@dataclass
class CaseResult:
    name: str
    passed: bool
    expected_path: str
    actual_path: str
    diff_count: int
    sample_diffs: List[DiffItem]
    error: str = ""


def _get_sheet(workbook: openpyxl.Workbook, sheet: str):
    if sheet == "active":
        return workbook.active
    return workbook[sheet]


def compare_workbooks(
    expected_file: Path,
    actual_file: Path,
    sheet_name: str = "active",
    max_diff_samples: int = 200,
) -> Tuple[bool, List[DiffItem], str]:
    if not expected_file.exists():
        return False, [], f"Expected file not found: {expected_file}"
    if not actual_file.exists():
        return False, [], f"Actual file not found: {actual_file}"

    expected_wb = None
    actual_wb = None
    try:
        expected_wb = openpyxl.load_workbook(expected_file, read_only=True, data_only=True)
        actual_wb = openpyxl.load_workbook(actual_file, read_only=True, data_only=True)
        expected_ws = _get_sheet(expected_wb, sheet_name)
        actual_ws = _get_sheet(actual_wb, sheet_name)

        max_row = max(expected_ws.max_row or 0, actual_ws.max_row or 0)
        max_col = max(expected_ws.max_column or 0, actual_ws.max_column or 0)

        diffs: List[DiffItem] = []
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                expected_value = expected_ws.cell(row=row_idx, column=col_idx).value
                actual_value = actual_ws.cell(row=row_idx, column=col_idx).value
                if expected_value != actual_value:
                    diffs.append(
                        DiffItem(
                            row=row_idx,
                            col=col_idx,
                            expected=expected_value,
                            actual=actual_value,
                        )
                    )
                    if len(diffs) >= max_diff_samples:
                        break
            if len(diffs) >= max_diff_samples:
                break
        return len(diffs) == 0, diffs, ""
    except Exception as exc:
        return False, [], str(exc)
    finally:
        if expected_wb is not None:
            expected_wb.close()
        if actual_wb is not None:
            actual_wb.close()


def run_manifest(manifest_path: Path) -> List[CaseResult]:
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    case_results: List[CaseResult] = []

    for case in manifest.get("cases", []):
        case_name = case["name"]
        expected = Path(case["expected"])
        actual = Path(case["actual"])
        sheet = case.get("sheet", "active")
        max_diff_samples = int(case.get("max_diff_samples", 200))

        passed, diffs, error = compare_workbooks(expected, actual, sheet, max_diff_samples)
        case_results.append(
            CaseResult(
                name=case_name,
                passed=passed,
                expected_path=str(expected),
                actual_path=str(actual),
                diff_count=len(diffs),
                sample_diffs=diffs[:20],
                error=error,
            )
        )

    return case_results


def generate_markdown_report(results: List[CaseResult], manifest_path: Path) -> str:
    passed = sum(1 for item in results if item.passed)
    failed = len(results) - passed

    lines = []
    lines.append("# Golden Regression Report")
    lines.append("")
    lines.append(f"- Time: {datetime.now().isoformat(timespec='seconds')}")
    lines.append(f"- Manifest: `{manifest_path}`")
    lines.append(f"- Total: {len(results)}")
    lines.append(f"- Passed: {passed}")
    lines.append(f"- Failed: {failed}")
    lines.append("")

    for result in results:
        status = "PASS" if result.passed else "FAIL"
        lines.append(f"## {result.name} [{status}]")
        lines.append(f"- Expected: `{result.expected_path}`")
        lines.append(f"- Actual: `{result.actual_path}`")
        lines.append(f"- Diff Count: {result.diff_count}")
        if result.error:
            lines.append(f"- Error: `{result.error}`")
        if result.sample_diffs:
            lines.append("")
            lines.append("| Row | Col | Expected | Actual |")
            lines.append("| --- | --- | --- | --- |")
            for diff in result.sample_diffs:
                lines.append(f"| {diff.row} | {diff.col} | {diff.expected!r} | {diff.actual!r} |")
        lines.append("")

    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description="Run golden workbook regression checks.")
    parser.add_argument(
        "--manifest",
        required=True,
        help="Path to manifest JSON file.",
    )
    parser.add_argument(
        "--report-dir",
        default="tests/golden/reports",
        help="Directory to write markdown report.",
    )
    args = parser.parse_args()

    manifest_path = Path(args.manifest)
    report_dir = Path(args.report_dir)
    report_dir.mkdir(parents=True, exist_ok=True)

    results = run_manifest(manifest_path)
    report_content = generate_markdown_report(results, manifest_path)
    report_name = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md"
    report_path = report_dir / report_name
    report_path.write_text(report_content, encoding="utf-8")

    print(f"Report written: {report_path}")
    if any(not result.passed for result in results):
        raise SystemExit(1)


if __name__ == "__main__":
    main()
