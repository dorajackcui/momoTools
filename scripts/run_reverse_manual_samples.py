from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SAMPLES_ROOT = ROOT / "tests" / "manual_data" / "reverse_stats_samples"
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from core.reverse_excel_processor import ReverseExcelProcessor


@dataclass(frozen=True)
class ManualCaseResult:
    name: str
    updated_count: int
    master_translation: str | None
    files_total: int
    files_succeeded: int
    files_failed: int
    error_codes: tuple[str, ...]
    logs: tuple[str, ...]


def write_workbook(path: Path, rows: list[list[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def reset_samples() -> None:
    write_workbook(
        SAMPLES_ROOT / "01_precedence" / "master.xlsx",
        [
            ["id", "key", "match", "translation"],
            ["", "K1", "M1", ""],
        ],
    )
    write_workbook(
        SAMPLES_ROOT / "01_precedence" / "targets" / "a.xlsx",
        [
            ["key", "match", "translation"],
            ["K1", "M1", "from_a"],
        ],
    )
    write_workbook(
        SAMPLES_ROOT / "01_precedence" / "targets" / "b.xlsx",
        [
            ["key", "match", "translation"],
            ["K1", "M1", "from_b"],
        ],
    )

    write_workbook(
        SAMPLES_ROOT / "02_read_failure" / "master.xlsx",
        [
            ["id", "key", "match", "translation"],
            ["", "K1", "M1", ""],
        ],
    )
    write_workbook(
        SAMPLES_ROOT / "02_read_failure" / "targets" / "a.xlsx",
        [
            ["key", "match", "translation"],
            ["K1", "M1", "from_good"],
        ],
    )
    bad_path = SAMPLES_ROOT / "02_read_failure" / "targets" / "b.xlsx"
    bad_path.parent.mkdir(parents=True, exist_ok=True)
    bad_path.write_text(
        "this is not a real excel workbook\nused to trigger E_TARGET_READ\n",
        encoding="utf-8",
    )


def read_cell(path: Path, row: int, col: int):
    workbook = openpyxl.load_workbook(path, data_only=True)
    try:
        return workbook.active.cell(row=row, column=col).value
    finally:
        workbook.close()


def run_case(case_name: str) -> ManualCaseResult:
    sample_root = SAMPLES_ROOT / case_name
    logs: list[str] = []
    processor = ReverseExcelProcessor(log_callback=logs.append)
    processor.set_master_file(str(sample_root / "master.xlsx"))
    processor.set_target_folder(str(sample_root / "targets"))
    updated_count = processor.process_files()

    return ManualCaseResult(
        name=case_name,
        updated_count=updated_count,
        master_translation=read_cell(sample_root / "master.xlsx", 2, 4),
        files_total=processor.stats.files_total,
        files_succeeded=processor.stats.files_succeeded,
        files_failed=processor.stats.files_failed,
        error_codes=tuple(event.code for event in processor.stats.errors),
        logs=tuple(logs),
    )


def print_result(result: ManualCaseResult) -> None:
    print(f"CASE {result.name}")
    print(f"  updated_count={result.updated_count}")
    print(f"  master_translation={result.master_translation!r}")
    print(
        "  stats="
        f"total:{result.files_total}, "
        f"succeeded:{result.files_succeeded}, "
        f"failed:{result.files_failed}, "
        f"errors:{len(result.error_codes)}"
    )
    if result.error_codes:
        print(f"  error_codes={list(result.error_codes)}")
    print("  logs:")
    for line in result.logs:
        print(f"    {line}")
    print()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Reset and run reverse-updater manual samples.",
    )
    parser.add_argument(
        "--no-reset",
        action="store_true",
        help="Keep the current sample files and run in-place.",
    )
    args = parser.parse_args()

    if not args.no_reset:
        reset_samples()

    for case_name in ("01_precedence", "02_read_failure"):
        print_result(run_case(case_name))


if __name__ == "__main__":
    main()
