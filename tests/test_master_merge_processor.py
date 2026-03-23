from datetime import date
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

from core.master_merge_processor import (
    CELL_WRITE_POLICY_FILL_BLANK_ONLY,
    CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
    KEY_ADMISSION_POLICY_ALLOW_NEW,
    KEY_ADMISSION_POLICY_EXISTING_ONLY,
    MasterMergeProcessor,
    PRIORITY_WINNER_POLICY_LAST_PROCESSED,
    ROW_KEY_POLICY_COMBINED,
    ROW_KEY_POLICY_KEY_ONLY,
)


def write_workbook(path: Path, rows: list[list[object]]) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def read_row(path: Path, row: int, max_col: int) -> list[object]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    try:
        worksheet = workbook.active
        return [worksheet.cell(row=row, column=col).value for col in range(1, max_col + 1)]
    finally:
        workbook.close()


def read_cell(path: Path, row: int, col: int):
    workbook = openpyxl.load_workbook(path, data_only=True)
    try:
        return workbook.active.cell(row=row, column=col).value
    finally:
        workbook.close()


def read_rows(path: Path) -> list[list[object]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    try:
        worksheet = workbook.active
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


class MasterMergeProcessorTestCase(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)

    def tearDown(self):
        self.temp_dir.cleanup()

    def _build_processor(self, master_path: Path, updates_dir: Path, files: list[Path]) -> MasterMergeProcessor:
        processor = MasterMergeProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_update_folder(str(updates_dir))
        processor.set_columns(0, 1)
        processor.set_priority_files([str(path) for path in files])
        return processor

    def test_merge_mode_append_only_skips_existing_keys_and_adds_new(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1", "v2"],
                ["K1", "M1", "", "MASTER_KEEP"],
            ],
        )

        high_path = updates_dir / "a.xlsx"
        low_path = updates_dir / "b.xlsx"

        write_workbook(
            high_path,
            [
                ["key", "match", "v1", "v2"],
                ["K1", "M1", "HIGH_V1", ""],
                ["K2", "M2", "NEW_HIGH", "HIGH_V2"],
            ],
        )
        write_workbook(
            low_path,
            [
                ["key", "match", "v1", "v2"],
                ["K1", "M1", "LOW_V1", "LOW_V2"],
                ["K2", "M2", "NEW_LOW", "LOW_V2"],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [high_path, low_path])
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_FILL_BLANK_ONLY,
            key_admission_policy=KEY_ADMISSION_POLICY_ALLOW_NEW,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 0)
        self.assertEqual(result.added_rows, 1)
        self.assertEqual(result.overwritten_cells, 0)
        self.assertEqual(result.filled_blank_cells, 0)
        self.assertEqual(result.skipped_new_keys, 0)
        self.assertEqual(result.unmatched_entries, 0)
        self.assertEqual(result.unmatched_report_path, "")

        self.assertEqual(read_row(master_path, 2, 4), ["K1", "M1", None, "MASTER_KEEP"])
        self.assertEqual(read_row(master_path, 3, 4), ["K2", "M2", "NEW_HIGH", "HIGH_V2"])

    def test_merge_key_only_appends_match_and_content_as_strings(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1", "v2"],
            ],
        )

        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1", "v2"],
                ["K2", "  M2  ", 0.25, date(2026, 3, 16)],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [source_path])
        processor.set_row_key_policy(ROW_KEY_POLICY_KEY_ONLY)
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_FILL_BLANK_ONLY,
            key_admission_policy=KEY_ADMISSION_POLICY_ALLOW_NEW,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.added_rows, 1)
        self.assertEqual(read_row(master_path, 2, 4), ["K2", "  M2  ", "0.25", "2026-03-16 00:00:00"])

    def test_merge_combined_key_treats_same_key_different_matches_as_distinct_rows(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1"],
            ],
        )

        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "FROM_M1"],
                ["K1", "M2", "FROM_M2"],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [source_path])
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_FILL_BLANK_ONLY,
            key_admission_policy=KEY_ADMISSION_POLICY_ALLOW_NEW,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.added_rows, 2)
        self.assertEqual(read_row(master_path, 2, 3), ["K1", "M1", "FROM_M1"])
        self.assertEqual(read_row(master_path, 3, 3), ["K1", "M2", "FROM_M2"])

    def test_merge_combined_key_duplicate_identity_keeps_first_priority_file_row(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1", "v2"],
            ],
        )

        first_path = updates_dir / "first.xlsx"
        second_path = updates_dir / "second.xlsx"
        write_workbook(
            first_path,
            [
                ["key", "match", "v1", "v2"],
                ["K1", "M1", "FIRST_V1", "FIRST_V2"],
            ],
        )
        write_workbook(
            second_path,
            [
                ["key", "match", "v1", "v2"],
                ["K1", "M1", "SECOND_V1", "SECOND_V2"],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [first_path, second_path])
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_FILL_BLANK_ONLY,
            key_admission_policy=KEY_ADMISSION_POLICY_ALLOW_NEW,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.added_rows, 1)
        self.assertEqual(read_row(master_path, 2, 4), ["K1", "M1", "FIRST_V1", "FIRST_V2"])

    def test_list_update_files_matches_internal_sorted_sources(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()
        write_workbook(master_path, [["key", "match", "v1"]])
        c_path = updates_dir / "c.xlsx"
        a_path = updates_dir / "a.xlsx"
        write_workbook(c_path, [["key", "match", "v1"]])
        write_workbook(a_path, [["key", "match", "v1"]])

        processor = MasterMergeProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_update_folder(str(updates_dir))

        self.assertEqual(
            processor.list_update_files(),
            [str(a_path), str(c_path)],
        )

    def test_update_master_mode_overwrites_non_blank_and_keeps_last_processed(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "MASTER_V"],
            ],
        )

        first_path = updates_dir / "first.xlsx"
        second_path = updates_dir / "second.xlsx"

        write_workbook(
            first_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "FROM_FIRST"],
                ["K2", "M2", "NEW_FIRST"],
            ],
        )
        write_workbook(
            second_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "FROM_SECOND"],
                ["K2", "M2", "NEW_SECOND"],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [first_path, second_path])
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_ALLOW_NEW,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 1)
        self.assertEqual(result.added_rows, 1)
        self.assertEqual(result.overwritten_cells, 1)
        self.assertEqual(result.filled_blank_cells, 0)
        self.assertEqual(result.skipped_new_keys, 0)
        self.assertEqual(result.unmatched_entries, 0)
        self.assertEqual(result.unmatched_report_path, "")

        self.assertEqual(read_row(master_path, 2, 3), ["K1", "M1", "FROM_SECOND"])
        self.assertEqual(read_row(master_path, 3, 3), ["K2", "M2", "NEW_SECOND"])

    def test_update_content_mode_skips_new_keys_and_updates_existing(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", ""],
            ],
        )

        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "UPDATE_EXISTING"],
                ["K3", "M3", "SHOULD_SKIP"],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [source_path])
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_EXISTING_ONLY,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 1)
        self.assertEqual(result.added_rows, 0)
        self.assertEqual(result.filled_blank_cells, 1)
        self.assertEqual(result.overwritten_cells, 0)
        self.assertEqual(result.skipped_new_keys, 1)
        self.assertEqual(result.unmatched_entries, 1)
        self.assertTrue(result.unmatched_report_path)
        self.assertTrue(Path(result.unmatched_report_path).exists())
        self.assertEqual(
            read_rows(Path(result.unmatched_report_path)),
            [
                ["key", "match", "source_file", "content_col_3"],
                ["K3", "M3", "source.xlsx", "SHOULD_SKIP"],
            ],
        )
        self.assertEqual(read_row(master_path, 2, 3), ["K1", "M1", "UPDATE_EXISTING"])

    def test_duplicate_master_keys_are_synchronized(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "A"],
                ["K1", "M1", "B"],
            ],
        )

        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "SYNC"],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [source_path])
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_ALLOW_NEW,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 2)
        self.assertEqual(result.overwritten_cells, 2)
        self.assertEqual(read_cell(master_path, 2, 3), "SYNC")
        self.assertEqual(read_cell(master_path, 3, 3), "SYNC")

    def test_merge_key_only_mode_append_only_does_not_touch_existing_rows(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1"],
                ["K1", "", ""],
            ],
        )
        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "S1"],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [source_path])
        processor.set_row_key_policy(ROW_KEY_POLICY_KEY_ONLY)
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_FILL_BLANK_ONLY,
            key_admission_policy=KEY_ADMISSION_POLICY_ALLOW_NEW,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 0)
        self.assertEqual(result.added_rows, 0)
        self.assertEqual(result.filled_blank_cells, 0)
        self.assertEqual(read_row(master_path, 2, 3), ["K1", None, None])

    def test_merge_key_only_mode_skips_new_rows_with_blank_match(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1"],
            ],
        )
        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1"],
                ["K1", "", "SKIP_ME"],
                ["K2", "M2", "APPEND_ME"],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [source_path])
        processor.set_row_key_policy(ROW_KEY_POLICY_KEY_ONLY)
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_FILL_BLANK_ONLY,
            key_admission_policy=KEY_ADMISSION_POLICY_ALLOW_NEW,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 0)
        self.assertEqual(result.added_rows, 1)
        self.assertEqual(result.filled_blank_cells, 0)
        self.assertEqual(read_row(master_path, 2, 3), ["K2", "M2", "APPEND_ME"])

    def test_update_master_forces_key_only_and_overwrites_match(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1"],
                ["K1", "OLD_M", "OLD_V"],
            ],
        )
        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1"],
                ["K1", "NEW_M", "NEW_V"],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [source_path])
        processor.set_row_key_policy(ROW_KEY_POLICY_COMBINED)
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_ALLOW_NEW,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 2)
        self.assertEqual(result.overwritten_cells, 2)
        self.assertEqual(read_row(master_path, 2, 3), ["K1", "NEW_M", "NEW_V"])

    def test_update_master_stringifies_match_and_content_values(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1", "v2"],
                ["K1", "OLD_M", "OLD_V1", "OLD_V2"],
            ],
        )
        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1", "v2"],
                ["K1", "  NEW_M  ", 0.25, date(2026, 3, 16)],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [source_path])
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_ALLOW_NEW,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 3)
        self.assertEqual(result.overwritten_cells, 3)
        self.assertEqual(read_row(master_path, 2, 4), ["K1", "  NEW_M  ", "0.25", "2026-03-16 00:00:00"])

    def test_update_content_forces_combined_key(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "OLD_V"],
            ],
        )
        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1"],
                ["K1", "M2", "NEW_V"],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [source_path])
        processor.set_row_key_policy(ROW_KEY_POLICY_KEY_ONLY)
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_EXISTING_ONLY,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 0)
        self.assertEqual(result.skipped_new_keys, 1)
        self.assertEqual(result.unmatched_entries, 1)
        self.assertTrue(Path(result.unmatched_report_path).exists())
        self.assertEqual(
            read_rows(Path(result.unmatched_report_path)),
            [
                ["key", "match", "source_file", "content_col_3"],
                ["K1", "M2", "source.xlsx", "NEW_V"],
            ],
        )
        self.assertEqual(read_row(master_path, 2, 3), ["K1", "M1", "OLD_V"])

    def test_update_content_unmatched_report_keeps_one_row_per_unmatched_identity(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "OLD_V"],
            ],
        )

        first_path = updates_dir / "a.xlsx"
        second_path = updates_dir / "b.xlsx"
        write_workbook(
            first_path,
            [
                ["key", "match", "v1"],
                ["K2", "M2", "MISS_1"],
                ["K2", "M3", "MISS_2"],
            ],
        )
        write_workbook(
            second_path,
            [
                ["key", "match", "v1"],
                ["K2", "M2", "MISS_1_LATER"],
                ["K3", "M4", "MISS_3"],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [first_path, second_path])
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_EXISTING_ONLY,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 0)
        self.assertEqual(result.skipped_new_keys, 3)
        self.assertEqual(result.unmatched_entries, 3)
        self.assertTrue(Path(result.unmatched_report_path).exists())
        self.assertEqual(
            read_rows(Path(result.unmatched_report_path)),
            [
                ["key", "match", "source_file", "content_col_3"],
                ["K2", "M2", "b.xlsx", "MISS_1_LATER"],
                ["K2", "M3", "a.xlsx", "MISS_2"],
                ["K3", "M4", "b.xlsx", "MISS_3"],
            ],
        )

    def test_update_content_always_generates_report_even_when_no_unmatched_entries(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "OLD_V"],
            ],
        )
        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "NEW_V"],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [source_path])
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_EXISTING_ONLY,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 1)
        self.assertEqual(result.skipped_new_keys, 0)
        self.assertEqual(result.unmatched_entries, 0)
        self.assertTrue(Path(result.unmatched_report_path).exists())
        self.assertEqual(
            read_rows(Path(result.unmatched_report_path)),
            [["key", "match", "source_file", "content_col_3"]],
        )

    def test_update_content_unmatched_report_expands_multiple_content_columns(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1", "v2"],
                ["K1", "M1", "OLD_V1", "OLD_V2"],
            ],
        )
        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1", "v2"],
                ["K2", "M2", "MISS_V1", "MISS_V2"],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [source_path])
        processor.set_columns(0, 1, 3)
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_EXISTING_ONLY,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.skipped_new_keys, 1)
        self.assertEqual(
            read_rows(Path(result.unmatched_report_path)),
            [
                ["key", "match", "source_file", "content_col_3", "content_col_4"],
                ["K2", "M2", "source.xlsx", "MISS_V1", "MISS_V2"],
            ],
        )

    def test_update_content_unmatched_report_stringifies_content_values(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1", "v2"],
                ["K1", "M1", "OLD_V1", "OLD_V2"],
            ],
        )
        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1", "v2"],
                ["K2", "M2", 0, date(2026, 3, 16)],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [source_path])
        processor.set_columns(0, 1, 3)
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_EXISTING_ONLY,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.skipped_new_keys, 1)
        self.assertEqual(
            read_rows(Path(result.unmatched_report_path)),
            [
                ["key", "match", "source_file", "content_col_3", "content_col_4"],
                ["K2", "M2", "source.xlsx", "0", "2026-03-16 00:00:00"],
            ],
        )

    def test_update_master_dense_rows_overwrite_untouched_cols_too(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1", "v2", "v3"],
                ["K1", "M1", "KEEP_V1", "OLD_V2", "KEEP_V3"],
            ],
        )
        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1", "v2", "v3"],
                ["K1", "M1", "", "NEW_V2", ""],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [source_path])
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_ALLOW_NEW,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 3)
        self.assertEqual(result.overwritten_cells, 1)
        self.assertEqual(result.filled_blank_cells, 0)
        self.assertEqual(read_row(master_path, 2, 5), ["K1", "M1", None, "NEW_V2", None])

    def test_update_master_dense_blank_values_clear_existing_cells(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1", "v2"],
                ["K1", "M1", "OLD_V1", "OLD_V2"],
            ],
        )
        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1", "v2"],
                ["K1", "M1", "", "NEW_V2"],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [source_path])
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_ALLOW_NEW,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 2)
        self.assertEqual(result.overwritten_cells, 1)
        self.assertEqual(result.filled_blank_cells, 0)
        self.assertEqual(read_row(master_path, 2, 4), ["K1", "M1", None, "NEW_V2"])

    def test_update_master_dense_duplicate_keys_use_last_processed_whole_row(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1", "v2"],
                ["K1", "OLD_M", "OLD_V1", "OLD_V2"],
            ],
        )

        first_path = updates_dir / "first.xlsx"
        second_path = updates_dir / "second.xlsx"
        write_workbook(
            first_path,
            [
                ["key", "match", "v1", "v2"],
                ["K1", "FIRST_M", "FIRST_V1", "FIRST_V2"],
            ],
        )
        write_workbook(
            second_path,
            [
                ["key", "match", "v1", "v2"],
                ["K1", "SECOND_M", "SECOND_V1", "SECOND_V2"],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [first_path, second_path])
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_ALLOW_NEW,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 3)
        self.assertEqual(result.overwritten_cells, 3)
        self.assertEqual(read_row(master_path, 2, 4), ["K1", "SECOND_M", "SECOND_V1", "SECOND_V2"])

    def test_update_content_sparse_blank_cells_do_not_clear_existing(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "OLD_V"],
            ],
        )
        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", ""],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [source_path])
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_EXISTING_ONLY,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 0)
        self.assertEqual(result.overwritten_cells, 0)
        self.assertEqual(result.filled_blank_cells, 0)
        self.assertEqual(read_row(master_path, 2, 3), ["K1", "M1", "OLD_V"])

    def test_update_content_combined_key_keeps_last_processed_values_for_touched_columns(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1", "v2"],
                ["K1", "M1", "OLD_V1", "OLD_V2"],
            ],
        )

        first_path = updates_dir / "first.xlsx"
        second_path = updates_dir / "second.xlsx"
        write_workbook(
            first_path,
            [
                ["key", "match", "v1", "v2"],
                ["K1", "M1", "FIRST_V1", ""],
            ],
        )
        write_workbook(
            second_path,
            [
                ["key", "match", "v1", "v2"],
                ["K1", "M1", "SECOND_V1", "SECOND_V2"],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [first_path, second_path])
        processor.set_columns(0, 1, 3)
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_EXISTING_ONLY,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 2)
        self.assertEqual(result.overwritten_cells, 2)
        self.assertEqual(result.skipped_new_keys, 0)
        self.assertEqual(read_row(master_path, 2, 4), ["K1", "M1", "SECOND_V1", "SECOND_V2"])

    def test_update_content_combined_key_does_not_cross_update_same_key_different_match(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "KEEP_M1"],
                ["K2", "M2", "OLD_M2"],
            ],
        )

        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1"],
                ["K1", "M2", "WRONG_MATCH"],
                ["K2", "M2", "RIGHT_MATCH"],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [source_path])
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_EXISTING_ONLY,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 1)
        self.assertEqual(result.skipped_new_keys, 1)
        self.assertEqual(result.unmatched_entries, 1)
        self.assertEqual(read_row(master_path, 2, 3), ["K1", "M1", "KEEP_M1"])
        self.assertEqual(read_row(master_path, 3, 3), ["K2", "M2", "RIGHT_MATCH"])
        self.assertEqual(
            read_rows(Path(result.unmatched_report_path)),
            [
                ["key", "match", "source_file", "content_col_3"],
                ["K1", "M2", "source.xlsx", "WRONG_MATCH"],
            ],
        )

    def test_update_content_sparse_stringifies_values_and_skips_whitespace_only(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1", "v2", "v3"],
                ["K1", "M1", "OLD_V1", "KEEP_V2", "OLD_V3"],
            ],
        )
        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1", "v2", "v3"],
                ["K1", "M1", 12.34, "   ", date(2026, 3, 16)],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [source_path])
        processor.set_columns(0, 1, 4)
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_EXISTING_ONLY,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 2)
        self.assertEqual(result.overwritten_cells, 2)
        self.assertEqual(result.filled_blank_cells, 0)
        self.assertEqual(
            read_row(master_path, 2, 5),
            ["K1", "M1", "12.34", "KEEP_V2", "2026-03-16 00:00:00"],
        )

    def test_update_master_noop_keeps_write_stage_zero_and_logs_perf(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "UNCHANGED"],
            ],
        )
        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "UNCHANGED"],
            ],
        )

        logs: list[str] = []
        processor = MasterMergeProcessor(log_callback=logs.append)
        processor.set_master_file(str(master_path))
        processor.set_update_folder(str(updates_dir))
        processor.set_columns(0, 1)
        processor.set_priority_files([str(source_path)])
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_ALLOW_NEW,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 0)
        self.assertEqual(result.added_rows, 0)
        perf_logs = [line for line in logs if line.startswith("Perf(Update Master):")]
        self.assertTrue(perf_logs)
        self.assertIn("layout_probe_used=", perf_logs[-1])
        self.assertIn("layout_probe=", perf_logs[-1])
        self.assertIn("open_master_rw_apply=0.00s", perf_logs[-1])
        self.assertIn("save_master=0.00s", perf_logs[-1])

    def test_update_content_logs_perf_summary(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "OLD_V"],
            ],
        )
        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "NEW_V"],
            ],
        )

        logs: list[str] = []
        processor = MasterMergeProcessor(log_callback=logs.append)
        processor.set_master_file(str(master_path))
        processor.set_update_folder(str(updates_dir))
        processor.set_columns(0, 1)
        processor.set_priority_files([str(source_path)])
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_EXISTING_ONLY,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        result = processor.process_files()

        self.assertEqual(result.updated_cells, 1)
        perf_logs = [line for line in logs if line.startswith("Perf(Update Content):")]
        self.assertTrue(perf_logs)
        self.assertIn("layout_probe_used=", perf_logs[-1])
        self.assertIn("layout_probe=", perf_logs[-1])

    def test_update_master_read_only_loads_disable_external_links(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "OLD_V"],
            ],
        )
        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "NEW_V"],
            ],
        )

        processor = self._build_processor(master_path, updates_dir, [source_path])
        processor.set_columns(0, 1, 2)
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_ALLOW_NEW,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        real_loader = openpyxl.load_workbook
        seen_load_kwargs: list[dict[str, object]] = []

        def tracking_loader(*args, **kwargs):
            seen_load_kwargs.append(dict(kwargs))
            return real_loader(*args, **kwargs)

        with patch(
            "core.kernel.excel_io.openpyxl.load_workbook",
            side_effect=tracking_loader,
        ):
            result = processor.process_files()

        self.assertEqual(result.updated_cells, 1)
        read_only_calls = [kwargs for kwargs in seen_load_kwargs if kwargs.get("read_only")]
        self.assertTrue(read_only_calls)
        self.assertTrue(all(kwargs.get("keep_links") is False for kwargs in read_only_calls))

        read_write_calls = [kwargs for kwargs in seen_load_kwargs if not kwargs.get("read_only")]
        self.assertTrue(read_write_calls)
        self.assertTrue(all("keep_links" not in kwargs for kwargs in read_write_calls))

    def test_update_master_skips_master_scan_when_no_valid_source_candidates(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "OLD_V"],
            ],
        )
        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1"],
                ["", "M2", "IGNORED"],
            ],
        )

        logs: list[str] = []
        processor = MasterMergeProcessor(log_callback=logs.append)
        processor.set_master_file(str(master_path))
        processor.set_update_folder(str(updates_dir))
        processor.set_columns(0, 1, 2)
        processor.set_priority_files([str(source_path)])
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_ALLOW_NEW,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        with patch(
            "core.master_update.executors.update_master.open_workbook",
            side_effect=AssertionError("master scan should be skipped"),
        ):
            result = processor.process_files()

        self.assertEqual(result.updated_cells, 0)
        self.assertEqual(result.added_rows, 0)
        self.assertEqual(result.merged_keys, 0)
        perf_logs = [line for line in logs if line.startswith("Perf(Update Master):")]
        self.assertTrue(perf_logs)
        self.assertIn("scan_master_ro=0.00s", perf_logs[-1])

    def test_update_content_skips_master_scan_when_no_valid_source_candidates(self):
        master_path = self.root / "master.xlsx"
        updates_dir = self.root / "updates"
        updates_dir.mkdir()

        write_workbook(
            master_path,
            [
                ["key", "match", "v1"],
                ["K1", "M1", "OLD_V"],
            ],
        )
        source_path = updates_dir / "source.xlsx"
        write_workbook(
            source_path,
            [
                ["key", "match", "v1"],
                ["K1", "", "IGNORED"],
            ],
        )

        logs: list[str] = []
        processor = MasterMergeProcessor(log_callback=logs.append)
        processor.set_master_file(str(master_path))
        processor.set_update_folder(str(updates_dir))
        processor.set_columns(0, 1, 2)
        processor.set_priority_files([str(source_path)])
        processor.set_policies(
            cell_write_policy=CELL_WRITE_POLICY_OVERWRITE_NON_BLANK,
            key_admission_policy=KEY_ADMISSION_POLICY_EXISTING_ONLY,
            priority_winner_policy=PRIORITY_WINNER_POLICY_LAST_PROCESSED,
        )

        with patch(
            "core.master_update.executors.update_master.open_workbook",
            side_effect=AssertionError("master scan should be skipped"),
        ):
            result = processor.process_files()

        self.assertEqual(result.updated_cells, 0)
        self.assertEqual(result.added_rows, 0)
        self.assertEqual(result.skipped_new_keys, 0)
        self.assertEqual(result.unmatched_entries, 0)
        self.assertEqual(result.merged_keys, 0)
        self.assertTrue(Path(result.unmatched_report_path).exists())
        self.assertEqual(
            read_rows(Path(result.unmatched_report_path)),
            [["key", "match", "source_file", "content_col_3"]],
        )
        perf_logs = [line for line in logs if line.startswith("Perf(Update Content):")]
        self.assertTrue(perf_logs)
        self.assertIn("scan_master_ro=0.00s", perf_logs[-1])


if __name__ == "__main__":
    unittest.main()
