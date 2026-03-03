import tempfile
import unittest
from pathlib import Path

import openpyxl

from core.excel_processor import ExcelProcessor
from core.kernel import is_blank_value


PRIVATE_ROOT = Path("tests/_private_data")
MASTER_PATH = PRIVATE_ROOT / "master" / "master.xlsx"
PACKAGES_ROOT = PRIVATE_ROOT / "packages"

LANG_SPECS = {
    "es": {
        "folder": "[es]",
        "master_content_col": 6,  # 0-based, column G
    },
    "fr": {
        "folder": "[fr]",
        "master_content_col": 5,  # 0-based, column F
    },
}


def _is_non_blank(value) -> bool:
    if value is None:
        return False
    return str(value).strip() != ""


def _read_target_translation(target_path: Path):
    workbook = openpyxl.load_workbook(target_path, data_only=True)
    try:
        return workbook.active.cell(row=2, column=3).value
    finally:
        workbook.close()


def _write_master_case(path: Path, key: str, match: str, content_col: int, content_value: str) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    max_col = max(content_col + 1, 7)
    headers = [f"col_{index}" for index in range(1, max_col + 1)]
    headers[1] = "Key"
    headers[2] = "MsgStr"
    worksheet.append(headers)

    row = [""] * max_col
    row[1] = key
    row[2] = match
    row[content_col] = content_value
    worksheet.append(row)

    workbook.save(path)
    workbook.close()


def _write_target_case(path: Path, key: str, match: str, translation: str) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Key", "MsgStr", "Translation"])
    worksheet.append([key, match, translation])
    workbook.save(path)
    workbook.close()


class PrivateDataSkipBlankWriteTestCase(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.samples = cls._discover_samples_or_skip()

    @classmethod
    def _discover_samples_or_skip(cls):
        if not MASTER_PATH.exists():
            raise unittest.SkipTest(f"Missing private master file: {MASTER_PATH}")
        if not PACKAGES_ROOT.exists():
            raise unittest.SkipTest(f"Missing private packages folder: {PACKAGES_ROOT}")

        try:
            workbook = openpyxl.load_workbook(MASTER_PATH, read_only=True, data_only=True)
            max_col = workbook.active.max_column
            workbook.close()
        except Exception as exc:
            raise unittest.SkipTest(f"Cannot read private master file: {exc}") from exc

        if max_col < 7:
            raise unittest.SkipTest(f"Private master file has insufficient columns: {max_col}")

        samples: dict[str, tuple[str, str]] = {}
        for lang, spec in LANG_SPECS.items():
            lang_root = PACKAGES_ROOT / spec["folder"]
            if not lang_root.exists():
                raise unittest.SkipTest(f"Missing language package folder: {lang_root}")
            files = sorted(lang_root.rglob("*.xlsx"), key=lambda p: str(p).lower())
            if not files:
                raise unittest.SkipTest(f"No xlsx files found under: {lang_root}")

            pair = cls._extract_first_valid_pair(files)
            if pair is None:
                raise unittest.SkipTest(f"No valid key/match row found under: {lang_root}")
            samples[lang] = pair

        return samples

    @staticmethod
    def _extract_first_valid_pair(file_paths: list[Path]):
        for path in file_paths:
            try:
                workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
                worksheet = workbook.active
                for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
                    key, match = row
                    if _is_non_blank(key) and _is_non_blank(match):
                        workbook.close()
                        return str(key), str(match)
                workbook.close()
            except Exception:
                continue
        return None

    def _run_single_case(self, lang: str, allow_blank_write: bool):
        key, match = self.samples[lang]
        content_col = LANG_SPECS[lang]["master_content_col"]

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_root = Path(temp_dir)
            master_path = temp_root / "master_case.xlsx"
            target_folder = temp_root / "targets"
            target_folder.mkdir(parents=True, exist_ok=True)
            target_path = target_folder / "target_case.xlsx"

            _write_master_case(
                master_path,
                key=key,
                match=match,
                content_col=content_col,
                content_value="",
            )
            _write_target_case(
                target_path,
                key=key,
                match=match,
                translation="KEEP_ME",
            )

            processor = ExcelProcessor(log_callback=lambda _msg: None)
            processor.set_master_file(str(master_path))
            processor.set_target_folder(str(target_folder))
            processor.set_target_column(0, 1, 2)
            processor.set_master_column(1, 2, content_col)
            processor.set_fill_blank_only(False)
            processor.set_allow_blank_write(allow_blank_write)
            processor.set_post_process_enabled(False)

            updated_count = processor.process_files()
            updated_value = _read_target_translation(target_path)

        return updated_count, updated_value

    def test_skip_blank_write_when_allow_blank_write_false(self):
        for lang in ("es", "fr"):
            with self.subTest(lang=lang):
                updated_count, updated_value = self._run_single_case(
                    lang=lang,
                    allow_blank_write=False,
                )
                self.assertEqual(updated_count, 0)
                self.assertEqual(updated_value, "KEEP_ME")

    def test_blank_write_can_clear_when_allow_blank_write_true(self):
        for lang in ("es", "fr"):
            with self.subTest(lang=lang):
                updated_count, updated_value = self._run_single_case(
                    lang=lang,
                    allow_blank_write=True,
                )
                self.assertEqual(updated_count, 1)
                self.assertTrue(is_blank_value(updated_value))


if __name__ == "__main__":
    unittest.main()
