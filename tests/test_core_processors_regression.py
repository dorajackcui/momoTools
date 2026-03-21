import shutil
import tempfile
import time
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

from core.deep_replace_processor import DeepReplaceProcessor
from core.excel_processor import ExcelProcessor
from core.kernel import get_stable_workers_cap, is_blank_value, run_parallel_map
from core.multi_column_processor import MultiColumnExcelProcessor
from core.reverse_excel_processor import ReverseExcelProcessor
from core.untranslated_stats_processor import UntranslatedStatsProcessor


def write_workbook(path: Path, rows: list[list[object]]) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def read_cell(path: Path, row: int, col: int):
    workbook = openpyxl.load_workbook(path, data_only=True)
    try:
        return workbook.active.cell(row=row, column=col).value
    finally:
        workbook.close()


class CoreProcessorsRegressionTestCase(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)

    def tearDown(self):
        self.temp_dir.cleanup()

    @staticmethod
    def _load_workbook_with_save_failure(real_loader, blocked_paths):
        normalized = {str(Path(path)) for path in blocked_paths}

        def loader(filename, *args, **kwargs):
            workbook = real_loader(filename, *args, **kwargs)
            if kwargs.get("read_only"):
                return workbook
            if str(Path(filename)) not in normalized:
                return workbook

            def broken_save(*_args, **_kwargs):
                raise PermissionError("locked file")

            workbook.save = broken_save
            return workbook

        return loader

    @staticmethod
    def _load_workbook_with_write_load_failure(real_loader, blocked_paths):
        normalized = {str(Path(path)) for path in blocked_paths}

        def loader(filename, *args, **kwargs):
            if kwargs.get("read_only"):
                return real_loader(filename, *args, **kwargs)
            if str(Path(filename)) in normalized:
                raise PermissionError("write-open denied")
            return real_loader(filename, *args, **kwargs)

        return loader

    def test_single_update_overwrite_and_fill_blank_only(self):
        master_path = self.root / "master.xlsx"
        target_folder = self.root / "targets_single"
        target_folder.mkdir()
        target_path = target_folder / "target.xlsx"

        write_workbook(
            master_path,
            [
                ["id", "key", "match", "value"],
                ["", "K1", "M1", "V1"],
            ],
        )
        write_workbook(
            target_path,
            [
                ["key", "match", "translation"],
                ["K1", "M1", ""],
                ["K1", "M1", "keep-me"],
            ],
        )

        processor = ExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_target_folder(str(target_folder))
        processor.set_post_process_enabled(False)
        processor.set_fill_blank_only(True)

        updated_count = processor.process_files()

        self.assertEqual(updated_count, 1)
        self.assertEqual(read_cell(target_path, 2, 3), "V1")
        self.assertEqual(read_cell(target_path, 3, 3), "keep-me")

    def test_single_update_requires_exact_combined_key_match(self):
        master_path = self.root / "master_exact_single.xlsx"
        target_folder = self.root / "targets_exact_single"
        target_folder.mkdir()
        target_path = target_folder / "target.xlsx"

        write_workbook(
            master_path,
            [
                ["id", "key", "match", "value"],
                ["", "K1", "M1", "V1"],
            ],
        )
        write_workbook(
            target_path,
            [
                ["key", "match", "translation"],
                ["K1", "M2", "keep-wrong-match"],
                ["K1", "M1", ""],
            ],
        )

        processor = ExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_target_folder(str(target_folder))
        processor.set_post_process_enabled(False)

        updated_count = processor.process_files()

        self.assertEqual(updated_count, 1)
        self.assertEqual(read_cell(target_path, 2, 3), "keep-wrong-match")
        self.assertEqual(read_cell(target_path, 3, 3), "V1")

    def test_single_processor_list_target_files_matches_processing_candidates(self):
        target_folder = self.root / "targets_list_single"
        target_folder.mkdir()
        first_path = target_folder / "a.xlsx"
        second_path = target_folder / "nested" / "b.xlsx"
        second_path.parent.mkdir()
        write_workbook(first_path, [["key", "match", "translation"]])
        write_workbook(second_path, [["key", "match", "translation"]])

        processor = ExcelProcessor(log_callback=lambda _msg: None)
        processor.set_target_folder(str(target_folder))

        self.assertEqual(
            processor.list_target_files(),
            processor._list_target_files_internal(),
        )

    def test_reverse_processor_list_target_files_matches_processing_order(self):
        target_folder = self.root / "targets_list_reverse"
        target_folder.mkdir()
        b_path = target_folder / "b.xlsx"
        a_path = target_folder / "a.xlsx"
        write_workbook(b_path, [["key", "match", "translation"]])
        write_workbook(a_path, [["key", "match", "translation"]])

        processor = ReverseExcelProcessor(log_callback=lambda _msg: None)
        processor.set_target_folder(str(target_folder))

        self.assertEqual(
            processor.list_target_files(),
            [str(a_path), str(b_path)],
        )

    def test_multi_column_update(self):
        master_path = self.root / "master_multi.xlsx"
        target_folder = self.root / "targets_multi"
        target_folder.mkdir()
        target_path = target_folder / "target.xlsx"

        write_workbook(
            master_path,
            [
                ["id", "key", "match", "meta", "v1", "v2"],
                ["", "K1", "M1", "", "A1", "A2"],
            ],
        )
        write_workbook(
            target_path,
            [
                ["id", "key", "match", "meta", "out1", "out2"],
                ["", "K1", "M1", "", "", "old"],
            ],
        )

        processor = MultiColumnExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_target_folder(str(target_folder))
        processor.set_column_count(2)
        processor.set_post_process_enabled(False)
        processor.set_fill_blank_only(False)

        updated_count = processor.process_files()

        self.assertEqual(updated_count, 2)
        self.assertEqual(read_cell(target_path, 2, 5), "A1")
        self.assertEqual(read_cell(target_path, 2, 6), "A2")

    def test_multi_column_update_requires_exact_combined_key_match(self):
        master_path = self.root / "master_multi_exact.xlsx"
        target_folder = self.root / "targets_multi_exact"
        target_folder.mkdir()
        target_path = target_folder / "target.xlsx"

        write_workbook(
            master_path,
            [
                ["id", "key", "match", "meta", "v1", "v2"],
                ["", "K1", "M1", "", "A1", "A2"],
            ],
        )
        write_workbook(
            target_path,
            [
                ["id", "key", "match", "meta", "out1", "out2"],
                ["", "K1", "M2", "", "WRONG_1", "WRONG_2"],
                ["", "K1", "M1", "", "OLD_1", "OLD_2"],
            ],
        )

        processor = MultiColumnExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_target_folder(str(target_folder))
        processor.set_column_count(2)
        processor.set_post_process_enabled(False)
        processor.set_fill_blank_only(False)

        updated_count = processor.process_files()

        self.assertEqual(updated_count, 2)
        self.assertEqual(read_cell(target_path, 2, 5), "WRONG_1")
        self.assertEqual(read_cell(target_path, 2, 6), "WRONG_2")
        self.assertEqual(read_cell(target_path, 3, 5), "A1")
        self.assertEqual(read_cell(target_path, 3, 6), "A2")

    def test_single_update_with_sparse_wide_columns(self):
        master_path = self.root / "master_sparse.xlsx"
        target_folder = self.root / "targets_sparse_single"
        target_folder.mkdir()
        target_path = target_folder / "target_sparse.xlsx"

        master_header = ["id", "key", "match"] + [""] * 17 + ["value"]
        master_row = ["", "K1", "M1"] + [""] * 17 + ["WIDE_V"]
        target_header = ["key", "match"] + [""] * 28 + ["translation"]
        target_row = ["K1", "M1"] + [""] * 28 + ["old"]

        write_workbook(master_path, [master_header, master_row])
        write_workbook(target_path, [target_header, target_row])

        processor = ExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_target_folder(str(target_folder))
        processor.set_master_column(1, 2, 20)
        processor.set_target_column(0, 1, 30)
        processor.set_post_process_enabled(False)
        processor.set_fill_blank_only(False)

        updated_count = processor.process_files()

        self.assertEqual(updated_count, 1)
        self.assertEqual(read_cell(target_path, 2, 31), "WIDE_V")

    def test_single_skip_blank_write_by_default(self):
        master_path = self.root / "master_blank_single.xlsx"
        target_folder = self.root / "targets_blank_single"
        target_folder.mkdir()
        target_path = target_folder / "target.xlsx"

        write_workbook(
            master_path,
            [
                ["id", "key", "match", "value"],
                ["", "K1", "M1", ""],
            ],
        )
        write_workbook(
            target_path,
            [
                ["key", "match", "translation"],
                ["K1", "M1", "keep-me"],
            ],
        )

        processor = ExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_target_folder(str(target_folder))
        processor.set_post_process_enabled(False)
        processor.set_fill_blank_only(False)

        updated_count = processor.process_files()

        self.assertEqual(updated_count, 0)
        self.assertEqual(read_cell(target_path, 2, 3), "keep-me")

    def test_single_allow_blank_write_can_clear_target(self):
        master_path = self.root / "master_blank_single_allow.xlsx"
        target_folder = self.root / "targets_blank_single_allow"
        target_folder.mkdir()
        target_path = target_folder / "target.xlsx"

        write_workbook(
            master_path,
            [
                ["id", "key", "match", "value"],
                ["", "K1", "M1", ""],
            ],
        )
        write_workbook(
            target_path,
            [
                ["key", "match", "translation"],
                ["K1", "M1", "keep-me"],
            ],
        )

        processor = ExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_target_folder(str(target_folder))
        processor.set_post_process_enabled(False)
        processor.set_fill_blank_only(False)
        processor.set_allow_blank_write(True)

        updated_count = processor.process_files()

        self.assertEqual(updated_count, 1)
        self.assertTrue(is_blank_value(read_cell(target_path, 2, 3)))

    def test_single_update_records_save_failure_context(self):
        master_path = self.root / "master_save_error.xlsx"
        target_folder = self.root / "targets_save_error"
        target_folder.mkdir()
        target_path = target_folder / "target.xlsx"

        write_workbook(
            master_path,
            [
                ["id", "key", "match", "value"],
                ["", "K1", "M1", "V1"],
            ],
        )
        write_workbook(
            target_path,
            [
                ["key", "match", "translation"],
                ["K1", "M1", ""],
            ],
        )

        processor = ExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_target_folder(str(target_folder))
        processor.set_post_process_enabled(False)

        real_loader = openpyxl.load_workbook
        with (
            patch("core.excel_processor.get_stable_workers_cap", return_value=1),
            patch(
                "core.kernel.excel_io.openpyxl.load_workbook",
                side_effect=self._load_workbook_with_save_failure(real_loader, [target_path]),
            ),
        ):
            updated_count = processor.process_files()

        self.assertEqual(updated_count, 0)
        self.assertEqual(processor.stats.files_failed, 1)
        self.assertEqual(processor.stats.files_succeeded, 0)
        self.assertEqual(len(processor.stats.errors), 1)
        error = processor.stats.errors[0]
        self.assertEqual(error.code, "E_TARGET_SAVE")
        self.assertIsInstance(error.exception, PermissionError)
        self.assertEqual(error.context["stage"], "save")
        self.assertEqual(error.context["update_count"], 1)
        self.assertEqual(error.context["sample_row"], 2)
        self.assertEqual(error.context["sample_col"], 3)

    def test_single_update_records_load_failure_context(self):
        master_path = self.root / "master_load_error.xlsx"
        target_folder = self.root / "targets_load_error"
        target_folder.mkdir()
        target_path = target_folder / "target.xlsx"

        write_workbook(
            master_path,
            [
                ["id", "key", "match", "value"],
                ["", "K1", "M1", "V1"],
            ],
        )
        write_workbook(
            target_path,
            [
                ["key", "match", "translation"],
                ["K1", "M1", ""],
            ],
        )

        processor = ExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_target_folder(str(target_folder))
        processor.set_post_process_enabled(False)

        real_loader = openpyxl.load_workbook
        with (
            patch("core.excel_processor.get_stable_workers_cap", return_value=1),
            patch(
                "core.kernel.excel_io.openpyxl.load_workbook",
                side_effect=self._load_workbook_with_write_load_failure(real_loader, [target_path]),
            ),
        ):
            updated_count = processor.process_files()

        self.assertEqual(updated_count, 0)
        self.assertEqual(processor.stats.files_failed, 1)
        self.assertEqual(processor.stats.files_succeeded, 0)
        self.assertEqual(len(processor.stats.errors), 1)
        error = processor.stats.errors[0]
        self.assertEqual(error.code, "E_TARGET_SAVE")
        self.assertIsInstance(error.exception, PermissionError)
        self.assertEqual(error.context["stage"], "load")
        self.assertEqual(error.context["update_count"], 1)
        self.assertEqual(error.context["sample_row"], 2)
        self.assertEqual(error.context["sample_col"], 3)

    def test_single_update_continues_after_one_file_save_failure(self):
        master_path = self.root / "master_partial_save_error.xlsx"
        target_folder = self.root / "targets_partial_save_error"
        target_folder.mkdir()
        good_target = target_folder / "good.xlsx"
        bad_target = target_folder / "bad.xlsx"

        write_workbook(
            master_path,
            [
                ["id", "key", "match", "value"],
                ["", "K1", "M1", "V1"],
            ],
        )
        target_rows = [
            ["key", "match", "translation"],
            ["K1", "M1", ""],
        ]
        write_workbook(good_target, target_rows)
        write_workbook(bad_target, target_rows)

        processor = ExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_target_folder(str(target_folder))
        processor.set_post_process_enabled(False)

        real_loader = openpyxl.load_workbook
        with (
            patch("core.excel_processor.get_stable_workers_cap", return_value=1),
            patch(
                "core.kernel.excel_io.openpyxl.load_workbook",
                side_effect=self._load_workbook_with_save_failure(real_loader, [bad_target]),
            ),
        ):
            updated_count = processor.process_files()

        self.assertEqual(updated_count, 1)
        self.assertEqual(read_cell(good_target, 2, 3), "V1")
        self.assertEqual(read_cell(bad_target, 2, 3), None)
        self.assertEqual(processor.stats.files_succeeded, 1)
        self.assertEqual(processor.stats.files_failed, 1)
        self.assertEqual(len(processor.stats.errors), 1)
        self.assertEqual(processor.stats.errors[0].code, "E_TARGET_SAVE")

    def test_multi_column_fill_blank_only_preserves_non_blank_cells(self):
        master_path = self.root / "master_multi_blank.xlsx"
        target_folder = self.root / "targets_multi_blank"
        target_folder.mkdir()
        target_path = target_folder / "target.xlsx"

        write_workbook(
            master_path,
            [
                ["id", "key", "match", "meta", "v1", "v2"],
                ["", "K1", "M1", "", "A1", "A2"],
            ],
        )
        write_workbook(
            target_path,
            [
                ["id", "key", "match", "meta", "out1", "out2"],
                ["", "K1", "M1", "", "", "keep"],
            ],
        )

        processor = MultiColumnExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_target_folder(str(target_folder))
        processor.set_column_count(2)
        processor.set_post_process_enabled(False)
        processor.set_fill_blank_only(True)

        updated_count = processor.process_files()

        self.assertEqual(updated_count, 1)
        self.assertEqual(read_cell(target_path, 2, 5), "A1")
        self.assertEqual(read_cell(target_path, 2, 6), "keep")

    def test_multi_column_update_records_save_failure_context(self):
        master_path = self.root / "master_multi_save_error.xlsx"
        target_folder = self.root / "targets_multi_save_error"
        target_folder.mkdir()
        target_path = target_folder / "target.xlsx"

        write_workbook(
            master_path,
            [
                ["id", "key", "match", "meta", "v1", "v2"],
                ["", "K1", "M1", "", "A1", "A2"],
            ],
        )
        write_workbook(
            target_path,
            [
                ["id", "key", "match", "meta", "out1", "out2"],
                ["", "K1", "M1", "", "", ""],
            ],
        )

        processor = MultiColumnExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_target_folder(str(target_folder))
        processor.set_column_count(2)
        processor.set_post_process_enabled(False)

        real_loader = openpyxl.load_workbook
        with (
            patch("core.multi_column_processor.get_stable_workers_cap", return_value=1),
            patch(
                "core.kernel.excel_io.openpyxl.load_workbook",
                side_effect=self._load_workbook_with_save_failure(real_loader, [target_path]),
            ),
        ):
            updated_count = processor.process_files()

        self.assertEqual(updated_count, 0)
        self.assertEqual(processor.stats.files_failed, 1)
        self.assertEqual(processor.stats.files_succeeded, 0)
        self.assertEqual(len(processor.stats.errors), 1)
        error = processor.stats.errors[0]
        self.assertEqual(error.code, "E_TARGET_SAVE")
        self.assertIsInstance(error.exception, PermissionError)
        self.assertEqual(error.context["stage"], "save")
        self.assertEqual(error.context["update_count"], 2)
        self.assertEqual(error.context["sample_row"], 2)
        self.assertEqual(error.context["sample_col"], 5)

    def test_multi_column_skip_blank_content_per_cell_by_default(self):
        master_path = self.root / "master_multi_blank_content.xlsx"
        target_folder = self.root / "targets_multi_blank_content"
        target_folder.mkdir()
        target_path = target_folder / "target.xlsx"

        write_workbook(
            master_path,
            [
                ["id", "key", "match", "meta", "v1", "v2"],
                ["", "K1", "M1", "", "A1", ""],
            ],
        )
        write_workbook(
            target_path,
            [
                ["id", "key", "match", "meta", "out1", "out2"],
                ["", "K1", "M1", "", "old1", "old2"],
            ],
        )

        processor = MultiColumnExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_target_folder(str(target_folder))
        processor.set_column_count(2)
        processor.set_post_process_enabled(False)
        processor.set_fill_blank_only(False)

        updated_count = processor.process_files()

        self.assertEqual(updated_count, 1)
        self.assertEqual(read_cell(target_path, 2, 5), "A1")
        self.assertEqual(read_cell(target_path, 2, 6), "old2")

    def test_reverse_update_has_deterministic_merge_precedence(self):
        master_path = self.root / "master_reverse.xlsx"
        target_folder = self.root / "targets_reverse"
        target_folder.mkdir()

        write_workbook(
            master_path,
            [
                ["id", "key", "match", "translation"],
                ["", "K1", "M1", ""],
            ],
        )

        write_workbook(
            target_folder / "a.xlsx",
            [
                ["key", "match", "translation"],
                ["K1", "M1", "from_a"],
            ],
        )
        write_workbook(
            target_folder / "b.xlsx",
            [
                ["key", "match", "translation"],
                ["K1", "M1", "from_b"],
            ],
        )

        processor = ReverseExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_target_folder(str(target_folder))

        updated_count = processor.process_files()

        self.assertEqual(updated_count, 1)
        self.assertEqual(read_cell(master_path, 2, 4), "from_b")
        self.assertEqual(processor.stats.files_succeeded, 2)
        self.assertEqual(processor.stats.files_failed, 0)

    def test_reverse_update_requires_exact_combined_key_match(self):
        master_path = self.root / "master_reverse_exact.xlsx"
        target_folder = self.root / "targets_reverse_exact"
        target_folder.mkdir()
        target_path = target_folder / "source.xlsx"

        write_workbook(
            master_path,
            [
                ["id", "key", "match", "translation"],
                ["", "K1", "M1", ""],
                ["", "K1", "M2", "keep-m2"],
            ],
        )
        write_workbook(
            target_path,
            [
                ["key", "match", "translation"],
                ["K1", "M1", "from_target"],
            ],
        )

        processor = ReverseExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_target_folder(str(target_folder))

        updated_count = processor.process_files()

        self.assertEqual(updated_count, 1)
        self.assertEqual(read_cell(master_path, 2, 4), "from_target")
        self.assertEqual(read_cell(master_path, 3, 4), "keep-m2")

    def test_reverse_read_failure_updates_stats_and_keeps_successful_updates(self):
        master_path = self.root / "master_reverse_read_error.xlsx"
        target_folder = self.root / "targets_reverse_read_error"
        target_folder.mkdir()
        good_path = target_folder / "a.xlsx"
        bad_path = target_folder / "b.xlsx"

        write_workbook(
            master_path,
            [
                ["id", "key", "match", "translation"],
                ["", "K1", "M1", ""],
            ],
        )
        write_workbook(
            good_path,
            [
                ["key", "match", "translation"],
                ["K1", "M1", "from_good"],
            ],
        )
        write_workbook(
            bad_path,
            [
                ["key", "match", "translation"],
                ["K1", "M1", "from_bad"],
            ],
        )

        processor = ReverseExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_target_folder(str(target_folder))

        real_loader = openpyxl.load_workbook

        def loader(filename, *args, **kwargs):
            if kwargs.get("read_only") and str(Path(filename)) == str(bad_path):
                raise PermissionError("read-open denied")
            return real_loader(filename, *args, **kwargs)

        with patch(
            "core.kernel.excel_io.openpyxl.load_workbook",
            side_effect=loader,
        ):
            updated_count = processor.process_files()

        self.assertEqual(updated_count, 1)
        self.assertEqual(read_cell(master_path, 2, 4), "from_good")
        self.assertEqual(processor.stats.files_total, 2)
        self.assertEqual(processor.stats.files_succeeded, 1)
        self.assertEqual(processor.stats.files_failed, 1)
        self.assertEqual(len(processor.stats.errors), 1)
        error = processor.stats.errors[0]
        self.assertEqual(error.code, "E_TARGET_READ")
        self.assertEqual(error.file_path, str(bad_path))
        self.assertEqual(error.context["file_name"], bad_path.name)
        self.assertIsInstance(error.exception, PermissionError)

    def test_reverse_update_with_sparse_columns(self):
        master_path = self.root / "master_reverse_sparse.xlsx"
        target_folder = self.root / "targets_reverse_sparse"
        target_folder.mkdir()
        target_path = target_folder / "sparse_target.xlsx"

        master_header = [""] * 26
        master_header[5] = "key"
        master_header[15] = "match"
        master_header[25] = "translation"
        master_row = [""] * 26
        master_row[5] = "K1"
        master_row[15] = "M1"

        target_header = [""] * 21
        target_header[0] = "key"
        target_header[10] = "match"
        target_header[20] = "translation"
        target_row = [""] * 21
        target_row[0] = "K1"
        target_row[10] = "M1"
        target_row[20] = "from_sparse"

        write_workbook(master_path, [master_header, master_row])
        write_workbook(target_path, [target_header, target_row])

        processor = ReverseExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_target_folder(str(target_folder))
        processor.set_target_columns(0, 10, 20)
        processor.set_master_columns(5, 15, 25)

        updated_count = processor.process_files()

        self.assertEqual(updated_count, 1)
        self.assertEqual(read_cell(master_path, 2, 26), "from_sparse")

    def test_reverse_skip_blank_write_by_default(self):
        master_path = self.root / "master_reverse_blank.xlsx"
        target_folder = self.root / "targets_reverse_blank"
        target_folder.mkdir()

        write_workbook(
            master_path,
            [
                ["id", "key", "match", "translation"],
                ["", "K1", "M1", "keep-me"],
            ],
        )
        write_workbook(
            target_folder / "a.xlsx",
            [
                ["key", "match", "translation"],
                ["K1", "M1", ""],
            ],
        )

        processor = ReverseExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_target_folder(str(target_folder))

        updated_count = processor.process_files()

        self.assertEqual(updated_count, 0)
        self.assertEqual(read_cell(master_path, 2, 4), "keep-me")

    def test_reverse_allow_blank_write_can_clear_master(self):
        master_path = self.root / "master_reverse_blank_allow.xlsx"
        target_folder = self.root / "targets_reverse_blank_allow"
        target_folder.mkdir()

        write_workbook(
            master_path,
            [
                ["id", "key", "match", "translation"],
                ["", "K1", "M1", "keep-me"],
            ],
        )
        write_workbook(
            target_folder / "a.xlsx",
            [
                ["key", "match", "translation"],
                ["K1", "M1", ""],
            ],
        )

        processor = ReverseExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file(str(master_path))
        processor.set_target_folder(str(target_folder))
        processor.set_allow_blank_write(True)

        updated_count = processor.process_files()

        self.assertEqual(updated_count, 1)
        self.assertTrue(is_blank_value(read_cell(master_path, 2, 4)))

    def test_stable_workers_cap_bounds(self):
        with patch("core.kernel.excel_io.os.cpu_count", return_value=1):
            self.assertEqual(get_stable_workers_cap(), 2)
        with patch("core.kernel.excel_io.os.cpu_count", return_value=64):
            self.assertEqual(get_stable_workers_cap(), 8)

    def test_processors_use_stable_workers_cap(self):
        single = ExcelProcessor(log_callback=lambda _msg: None)
        single.set_master_file("master.xlsx")
        single.set_target_folder("targets")
        single.set_post_process_enabled(False)
        with (
            patch("core.excel_processor.get_stable_workers_cap", return_value=5),
            patch.object(single, "_read_master_dict", return_value={}),
            patch("core.excel_processor.iter_excel_files", return_value=["a.xlsx", "b.xlsx"]),
            patch("core.excel_processor.process_files_in_parallel", return_value=0) as single_parallel,
        ):
            single.process_files()
        self.assertEqual(single_parallel.call_args.kwargs["max_workers_cap"], 5)

        multi = MultiColumnExcelProcessor(log_callback=lambda _msg: None)
        multi.set_master_file("master.xlsx")
        multi.set_target_folder("targets")
        multi.set_post_process_enabled(False)
        with (
            patch("core.multi_column_processor.get_stable_workers_cap", return_value=7),
            patch.object(multi, "_build_usecols", return_value=[1, 2, 4]),
            patch.object(multi, "_read_master_dataframe", return_value=[]),
            patch.object(multi, "_build_master_dict", return_value={}),
            patch("core.multi_column_processor.iter_excel_files", return_value=["a.xlsx"]),
            patch("core.multi_column_processor.process_files_in_parallel", return_value=0) as multi_parallel,
        ):
            multi.process_files()
        self.assertEqual(multi_parallel.call_args.kwargs["max_workers_cap"], 7)

        reverse = ReverseExcelProcessor(log_callback=lambda _msg: None)
        reverse.set_master_file("master.xlsx")
        reverse.set_target_folder("targets")
        with (
            patch("core.reverse_excel_processor.get_stable_workers_cap", return_value=6),
            patch("core.reverse_excel_processor.iter_excel_files", return_value=["a.xlsx", "b.xlsx"]),
            patch("core.reverse_excel_processor.run_parallel_map", return_value=[]) as reverse_parallel,
            patch.object(reverse, "_update_master_file", return_value=0),
        ):
            reverse.process_files()
        self.assertEqual(reverse_parallel.call_args.kwargs["max_workers_cap"], 6)

    def test_single_post_process_only_updated_files(self):
        processor = ExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file("master.xlsx")
        processor.set_target_folder("targets")
        processor.set_post_process_enabled(True)

        file_paths = ["a.xlsx", "b.xlsx", "c.xlsx"]
        updates_by_file = {
            "a.xlsx": 1,
            "b.xlsx": 0,
            "c.xlsx": 2,
        }

        def fake_parallel(paths, worker, max_workers_cap):
            return sum(worker(path) for path in paths)

        def fake_process(file_path, _master_dict):
            return updates_by_file[file_path]

        with (
            patch.object(processor, "_read_master_dict", return_value={}),
            patch("core.excel_processor.iter_excel_files", return_value=file_paths),
            patch.object(processor, "_process_single_file", side_effect=fake_process),
            patch("core.excel_processor.process_files_in_parallel", side_effect=fake_parallel),
            patch.object(processor, "_post_process") as post_process_mock,
        ):
            updated_count = processor.process_files()

        self.assertEqual(updated_count, 3)
        post_process_mock.assert_called_once_with(["a.xlsx", "c.xlsx"])

    def test_multi_post_process_only_updated_files(self):
        processor = MultiColumnExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file("master.xlsx")
        processor.set_target_folder("targets")
        processor.set_post_process_enabled(True)

        file_paths = ["a.xlsx", "b.xlsx", "c.xlsx"]
        updates_by_file = {
            "a.xlsx": 0,
            "b.xlsx": 4,
            "c.xlsx": 0,
        }

        def fake_parallel(paths, worker, max_workers_cap):
            return sum(worker(path) for path in paths)

        def fake_process(file_path, _master_dict):
            return updates_by_file[file_path]

        with (
            patch.object(processor, "_build_usecols", return_value=[0, 1, 2]),
            patch.object(processor, "_read_master_dataframe", return_value=[]),
            patch.object(processor, "_build_master_dict", return_value={}),
            patch("core.multi_column_processor.iter_excel_files", return_value=file_paths),
            patch.object(processor, "_process_single_file", side_effect=fake_process),
            patch("core.multi_column_processor.process_files_in_parallel", side_effect=fake_parallel),
            patch.object(processor, "_post_process") as post_process_mock,
        ):
            updated_count = processor.process_files()

        self.assertEqual(updated_count, 4)
        post_process_mock.assert_called_once_with(["b.xlsx"])

    def test_single_post_process_skipped_when_no_updates(self):
        processor = ExcelProcessor(log_callback=lambda _msg: None)
        processor.set_master_file("master.xlsx")
        processor.set_target_folder("targets")
        processor.set_post_process_enabled(True)

        file_paths = ["a.xlsx", "b.xlsx"]

        def fake_parallel(paths, worker, max_workers_cap):
            return sum(worker(path) for path in paths)

        with (
            patch.object(processor, "_read_master_dict", return_value={}),
            patch("core.excel_processor.iter_excel_files", return_value=file_paths),
            patch.object(processor, "_process_single_file", return_value=0),
            patch("core.excel_processor.process_files_in_parallel", side_effect=fake_parallel),
            patch.object(processor, "_post_process") as post_process_mock,
        ):
            updated_count = processor.process_files()

        self.assertEqual(updated_count, 0)
        post_process_mock.assert_not_called()

    def test_untranslated_stats_and_export(self):
        target_folder = self.root / "stats_targets"
        target_folder.mkdir()
        source_path = target_folder / "stats.xlsx"
        output_path = self.root / "stats_output.xlsx"

        write_workbook(
            source_path,
            [
                ["id", "source", "translation"],
                [1, "hello world", ""],
                [2, "cat", "done"],
            ],
        )

        processor = UntranslatedStatsProcessor(log_callback=lambda _msg: None)
        processor.set_target_folder(str(target_folder))
        processor.set_stats_mode("english_words")
        processor.set_columns(1, 2)

        results = processor.process_files()

        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]["untranslated_chars"], 2)
        self.assertEqual(results[0]["untranslated_rows"], 1)
        self.assertEqual(results[0]["total_chars"], 3)
        self.assertEqual(results[0]["total_rows"], 2)
        self.assertTrue(processor.export_to_excel(str(output_path)))
        self.assertTrue(output_path.exists())

    def test_deep_replace_rolls_back_when_copy_fails(self):
        source_folder = self.root / "replace_source"
        target_folder = self.root / "replace_target"
        source_folder.mkdir()
        target_folder.mkdir()

        source_file = source_folder / "sample.xlsx"
        target_file = target_folder / "sample.xlsx"
        source_file.write_text("new-content", encoding="utf-8")
        target_file.write_text("old-content", encoding="utf-8")

        processor = DeepReplaceProcessor(log_callback=lambda _msg: None)
        processor.set_source_folder(str(source_folder))
        processor.set_target_folder(str(target_folder))

        original_copy2 = shutil.copy2
        call_count = {"value": 0}

        def flaky_copy2(src, dst, *args, **kwargs):
            call_count["value"] += 1
            if call_count["value"] == 2:
                raise OSError("simulated copy failure")
            return original_copy2(src, dst, *args, **kwargs)

        with patch("core.deep_replace_processor.shutil.copy2", side_effect=flaky_copy2):
            processed_files = processor.process_files()

        self.assertEqual(processed_files, 0)
        self.assertEqual(target_file.read_text(encoding="utf-8"), "old-content")
        self.assertFalse(Path(f"{target_file}.bak").exists())
        self.assertEqual(len(processor.stats.errors), 1)

    def test_deep_replace_list_target_files_lists_target_excel_candidates(self):
        target_folder = self.root / "replace_target_listing"
        nested_folder = target_folder / "nested"
        target_folder.mkdir()
        nested_folder.mkdir()
        first_path = target_folder / "sample.xlsx"
        second_path = nested_folder / "nested.xlsx"
        first_path.write_text("placeholder", encoding="utf-8")
        second_path.write_text("placeholder", encoding="utf-8")

        processor = DeepReplaceProcessor(log_callback=lambda _msg: None)
        processor.set_target_folder(str(target_folder))

        self.assertEqual(
            sorted(processor.list_target_files()),
            sorted([str(first_path), str(second_path)]),
        )

    def test_parallel_map_preserves_input_order(self):
        items = [1, 2, 3]

        def worker(item):
            time.sleep((4 - item) * 0.01)
            return item * 10

        results = run_parallel_map(items, worker, max_workers_cap=3)
        self.assertEqual(results, [10, 20, 30])


if __name__ == "__main__":
    unittest.main()
